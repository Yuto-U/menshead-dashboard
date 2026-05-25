"""Excelファイル別のローダー。

各ローダーは pandas DataFrame を返し、副作用（DB書き込み）は持たない。
DB書き込みは pipeline.py で一括処理する。
"""
from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

from etl.normalize import (
    hash_phone,
    normalize_cast_name,
    normalize_store,
    normalize_text,
    parse_priority,
    safe_float,
    safe_int,
    yymm_to_year_month,
)

STORE_SHEET_PATTERN = re.compile(r"^(上野|銀座|新宿)(\d{4})$")
TOTAL_SHEET_PATTERN = re.compile(r"^総合KPI(\d{4})$")
STORE_ID_MAP = {"新宿": 1, "銀座": 2, "上野": 3}

# 列ヘッダー → 内部キー のマッピング候補（表記揺れを吸収）
SALES_HEADERS = ["総売", "総合売上", "売上"]
CASE_COUNT_HEADERS = ["総合件数", "案件数", "件数"]
REPEAT_RATE_HEADERS = ["総合リピート", "リピート率"]

# コース列のヘッダー命名規則
COURSE_COL_DEFS = [
    # (course_id, new_or_repeat, count_header_candidates, amount_header_candidates)
    (1, "new", ["新規60案件数", "新規60件数"], ["新規60金額"]),
    (2, "new", ["新規90案件数", "新規90件数"], ["新規90金額"]),
    (3, "new", ["新規120案件数", "新規120件数"], ["新規120金額"]),
    (4, "new", ["新規オイル90案件数"], ["新規オイル90金額"]),
    (5, "new", ["新規オイル120案件数"], ["新規オイル120金額"]),
    (6, "new", ["新規クリーム案件数", "クリーム単体案件数"], ["新規クリーム金額", "クリーム単体金額"]),
    (1, "repeat", ["リピート60案件数", "リピート60件数"], ["リピート60金額"]),
    (2, "repeat", ["リピート90案件数", "リピート90件数"], ["リピート90金額"]),
    (3, "repeat", ["リピート120案件数", "リピート120件数"], ["リピート120金額"]),
    (4, "repeat", ["リピートオイル90案件数"], ["リピートオイル90金額"]),
    (5, "repeat", ["リピートオイル120案件数"], ["リピートオイル120金額"]),
]


def _header_map(ws) -> dict[str, int]:
    """1行目をヘッダーとして、ヘッダー文字列→0-based列番号 を返す。"""
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    result: dict[str, int] = {}
    for idx, value in enumerate(header):
        key = normalize_text(value)
        if key:
            result[key] = idx
    return result


def _first_existing_col(header_map: dict[str, int], candidates: list[str]) -> Optional[int]:
    for c in candidates:
        idx = header_map.get(c)
        if idx is not None:
            return idx
    return None


def _row_value(row: tuple, idx: Optional[int]):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def load_store_kpi(file_path: Path) -> dict[str, pd.DataFrame]:
    """`ヘッド店舗KPI.xlsx` → fact_daily_sales, fact_course_daily を抽出。"""
    file_path = Path(file_path)
    wb = load_workbook(file_path, data_only=True, read_only=True)

    daily_rows: list[dict] = []
    course_rows: list[dict] = []

    for sheet_name in wb.sheetnames:
        m = STORE_SHEET_PATTERN.match(sheet_name)
        if not m:
            continue
        store_name = m.group(1)
        store_id = STORE_ID_MAP[store_name]

        ws = wb[sheet_name]
        header_map = _header_map(ws)

        sales_col = _first_existing_col(header_map, SALES_HEADERS)
        case_col = _first_existing_col(header_map, CASE_COUNT_HEADERS)
        repeat_col = _first_existing_col(header_map, REPEAT_RATE_HEADERS)

        course_specs = []
        for course_id, nor, count_keys, amount_keys in COURSE_COL_DEFS:
            c_idx = _first_existing_col(header_map, count_keys)
            a_idx = _first_existing_col(header_map, amount_keys)
            if c_idx is not None or a_idx is not None:
                course_specs.append((course_id, nor, c_idx, a_idx))

        # 行5以降が日次データ。row 1-4 はヘッダー/月合計/先月/着地予想
        for row in ws.iter_rows(min_row=5, values_only=True):
            if not row:
                continue
            first = row[0]
            if not isinstance(first, datetime):
                continue
            date_value = first.date()

            sales = safe_int(_row_value(row, sales_col))
            if not sales:
                continue

            case_count = safe_int(_row_value(row, case_col)) or 0
            repeat_rate = safe_float(_row_value(row, repeat_col))
            unit_price = int(sales / case_count) if case_count else None

            new_count_sum = 0
            repeat_count_sum = 0
            for course_id, nor, c_idx, a_idx in course_specs:
                count = safe_int(_row_value(row, c_idx)) or 0
                amount = safe_int(_row_value(row, a_idx)) or 0
                if count or amount:
                    course_rows.append({
                        "date": date_value,
                        "store_id": store_id,
                        "course_id": course_id,
                        "new_or_repeat": nor,
                        "case_count": count,
                        "sales": amount,
                        "source_file": file_path.name,
                    })
                if nor == "new":
                    new_count_sum += count
                else:
                    repeat_count_sum += count

            daily_rows.append({
                "date": date_value,
                "store_id": store_id,
                "sales": sales,
                "case_count": case_count or None,
                "unit_price": unit_price,
                "gross_profit": None,
                "cost": None,
                "new_count": new_count_sum or None,
                "repeat_count": repeat_count_sum or None,
                "repeat_rate": repeat_rate,
                "source_file": file_path.name,
            })

    wb.close()

    fact_daily = pd.DataFrame(daily_rows).drop_duplicates(subset=["date", "store_id"], keep="last")
    fact_course = pd.DataFrame(course_rows).drop_duplicates(
        subset=["date", "store_id", "course_id", "new_or_repeat"], keep="last"
    )
    return {
        "fact_daily_sales": fact_daily,
        "fact_course_daily": fact_course,
    }


# ---------- 他のloaderはPhase2以降で実装 ----------

def load_cast_master(file_path: Path) -> dict[str, pd.DataFrame]:
    """`ヘッドスパニスト一覧管理表.xlsx` → dim_cast_seed / fact_recruiting_monthly。"""
    file_path = Path(file_path)
    wb = load_workbook(file_path, data_only=True, read_only=True)
    out: dict[str, pd.DataFrame] = {}

    # --- dim_cast_seed: 「ヘッド」シートからキャストマスタを抽出 ---
    if "ヘッド" in wb.sheetnames:
        ws = wb["ヘッド"]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if rows:
            header = [normalize_text(v) for v in rows[0]]

            def col(*cands: str) -> Optional[int]:
                for c in cands:
                    if c in header:
                        return header.index(c)
                return None

            c_name = col("店舗名", "スパニスト名", "源氏名", "名前")
            c_haken = col("派遣名", "本名")
            c_area = col("固定エリア", "店舗", "エリア")
            c_status = col("ステータス", "状態")
            c_hire = col("採用日", "入店日", "デビュー日")
            c_priority = col("優先度")
            c_min_h = col("最低時間")

            seed: list[dict] = []
            for r in rows[1:]:
                if c_name is None or c_name >= len(r):
                    break
                nm = normalize_cast_name(r[c_name]) if r[c_name] else None
                if not nm:
                    continue
                prio_raw = None
                prio_num = None
                if c_priority is not None and c_priority < len(r):
                    prio_raw = normalize_text(r[c_priority])
                    prio_num = parse_priority(r[c_priority])
                seed.append({
                    "cast_name": nm,
                    "haken_name": normalize_text(r[c_haken]) if c_haken is not None and c_haken < len(r) else None,
                    "fixed_area": normalize_text(r[c_area]) if c_area is not None and c_area < len(r) else None,
                    "status": normalize_text(r[c_status]) if c_status is not None and c_status < len(r) else None,
                    "hire_date": r[c_hire] if c_hire is not None and c_hire < len(r) else None,
                    "priority": prio_num,
                    "priority_raw": prio_raw,
                    "min_hours": normalize_text(r[c_min_h]) if c_min_h is not None and c_min_h < len(r) else None,
                })
            if seed:
                out["dim_cast_seed"] = pd.DataFrame(seed).drop_duplicates(subset=["cast_name"], keep="last")

    wb.close()
    return out


CAST_EVAL_SHEET_PATTERN = re.compile(r"^(\d{4})$")


SURVEY_SCORE_MAP = {
    "良": 3, "派遣良": 3, "店舗良": 3,
    "普": 2, "派遣普": 2, "店舗普": 2,
    "悪": 1, "派遣悪": 1, "店舗悪": 1,
}


def _survey_score(text: Optional[str]) -> Optional[int]:
    if not text:
        return None
    text = normalize_text(text)
    if not text:
        return None
    if text in SURVEY_SCORE_MAP:
        return SURVEY_SCORE_MAP[text]
    for key, val in SURVEY_SCORE_MAP.items():
        if key in text:
            return val
    return None


def load_cast_evaluation(file_path: Path) -> dict[str, pd.DataFrame]:
    """`aoスパニスト評価シート.xlsx` → dim_cast_seed + fact_cast_monthly。

    月別シート（YYMM）のヘッダ行はr4、データはr5以降。
    スクショで要求された全項目を取り込む（店舗側集計＋total＋アンケート＋備考）。
    """
    file_path = Path(file_path)
    wb = load_workbook(file_path, data_only=True, read_only=True)

    cast_master: dict[str, dict] = {}
    fact_rows: list[dict] = []

    for sn in wb.sheetnames:
        m = CAST_EVAL_SHEET_PATTERN.match(sn)
        if not m:
            continue
        ym = yymm_to_year_month(m.group(1))
        if not ym:
            continue
        ws = wb[sn]
        all_rows = list(ws.iter_rows(min_row=1, max_row=300, values_only=True))
        if len(all_rows) < 5:
            continue
        header_row = all_rows[3]  # r4
        headers = [normalize_text(v) if v else "" for v in header_row]

        # ヘッダ内の改行・スペースを無視してマッチ（「店舗\n案件数」≡「店舗 案件数」≡「店舗案件数」）
        norm_headers = [re.sub(r"\s+", "", h) for h in headers]

        def col_idx(name: str, occurrence: int = 0) -> Optional[int]:
            target = re.sub(r"\s+", "", name)
            found = 0
            for i, h in enumerate(norm_headers):
                if h == target:
                    if found == occurrence:
                        return i
                    found += 1
            return None

        def col(*cands: str) -> Optional[int]:
            for c in cands:
                idx = col_idx(c)
                if idx is not None:
                    return idx
            return None

        c_cast = col("スパニスト")
        c_haken = col("派遣名")
        c_store_label = col("店舗名")
        c_work_days = col("稼働日数")
        c_contract_h = col("契約時間")
        if c_contract_h is None:
            c_contract_h = 21  # c22 固定フォールバック（ヘッダに無いが値はc22に入る運用）
        c_work_h = col("稼働時間")
        c_priority = col("優先度")
        c_guarantee = col("保証額")
        c_staff = col("担当スタッフ")
        c_survey = col("アンケート")
        c_note = col("備考")
        c_total_main_nom = col("合計\n本指名数", "合計本指名数")

        # 店舗側（c34-c41 ブロック）
        c_store_case = col("店舗\n案件数", "店舗案件数")
        c_store_main_rate = col("店舗\n本指名率", "店舗本指名率")
        c_store_main_nom = col("店舗\n（本指名）", "店舗（本指名）")
        c_store_sales = col("店舗\n売上", "店舗売上")
        c_store_reward = col("店舗\n報酬", "店舗報酬")
        c_store_gp = col("店舗\n粗利", "店舗粗利")
        c_store_nom = col("店舗\n指名", "店舗指名")
        c_store_nom_rate = col("店舗\n指名率", "店舗指名率")

        # total列
        c_total_sales = col("total売上")
        c_total_reward = col("total報酬")
        c_total_gp = col("total粗利")
        c_total_case = col("total\n案件数", "total案件数")

        if c_cast is None:
            continue

        for r in all_rows[4:]:
            if c_cast >= len(r) or not r[c_cast]:
                continue
            cast_name = normalize_cast_name(r[c_cast])
            if not cast_name:
                continue

            def get(idx: Optional[int]):
                if idx is None or idx >= len(r):
                    return None
                return r[idx]

            haken_name = normalize_text(get(c_haken))
            # 「店舗名」列は今は源氏名と同一（旧運用の名残）→ dim_cast.fixed_area には入れない
            work_days = safe_int(get(c_work_days))
            contract_h = normalize_text(get(c_contract_h))
            work_h = safe_float(get(c_work_h))
            priority_raw = normalize_text(get(c_priority))
            guarantee = safe_int(get(c_guarantee))
            staff = normalize_text(get(c_staff))
            survey_text = normalize_text(get(c_survey))
            survey_score = _survey_score(survey_text)
            note = normalize_text(get(c_note))
            total_main = safe_int(get(c_total_main_nom))

            store_case = safe_int(get(c_store_case)) or 0
            store_main_rate = safe_float(get(c_store_main_rate))
            store_main_nom = safe_int(get(c_store_main_nom)) or 0
            store_sales = safe_int(get(c_store_sales)) or 0
            store_reward = safe_int(get(c_store_reward)) or 0
            store_gp = safe_int(get(c_store_gp)) or 0
            store_nom = safe_int(get(c_store_nom)) or 0
            store_nom_rate = safe_float(get(c_store_nom_rate))

            total_sales = safe_int(get(c_total_sales)) or store_sales
            total_reward = safe_int(get(c_total_reward)) or store_reward
            total_gp = safe_int(get(c_total_gp)) or store_gp
            total_case = safe_int(get(c_total_case)) or store_case

            # 完全空行はスキップ
            if total_sales == 0 and total_case == 0 and not survey_text and not note:
                continue

            cast_master[cast_name] = {
                "cast_name": cast_name,
                "haken_name": haken_name,
                "fixed_area": None,
                "status": None,
                "hire_date": None,
            }

            fact_rows.append({
                "year_month": ym,
                "cast_name": cast_name,
                # 既存総合（後方互換）
                "sales": total_sales,
                "reward": total_reward,
                "gross_profit": total_gp,
                "case_count": total_case,
                "work_hours": work_h,
                "store_nomination_rate": store_nom_rate,
                "store_main_nomination_rate": store_main_rate,
                "haken_nomination_rate": None,
                "haken_main_nomination_rate": None,
                "survey_score": survey_score,
                # 拡張列
                "work_days": work_days,
                "contract_hours": contract_h,
                "priority_raw": priority_raw,
                "sales_store": store_sales,
                "reward_store": store_reward,
                "gross_profit_store": store_gp,
                "case_count_store": store_case,
                "nomination_store": store_nom,
                "main_nomination_store": store_main_nom,
                "nomination_rate_store2": store_nom_rate,
                "main_nomination_total": total_main,
                "total_sales": total_sales,
                "total_reward": total_reward,
                "total_gross_profit": total_gp,
                "total_case_count": total_case,
                "guarantee_amount": guarantee,
                "staff_name": staff,
                "survey_text": survey_text,
                "note": note,
                "source_file": file_path.name,
            })

    wb.close()
    result: dict[str, pd.DataFrame] = {}
    if cast_master:
        result["dim_cast_seed"] = pd.DataFrame(list(cast_master.values()))
    if fact_rows:
        result["fact_cast_monthly"] = pd.DataFrame(fact_rows)
    return result


def load_daily_report(file_path: Path) -> dict[str, pd.DataFrame]:
    """`ao合計日報{YYMM}.xlsx` から下記を抽出（4ファイル運用のメインソース）：

    - fact_daily_sales（店舗別合計シートから店舗×日の売上/報酬/純利益）
    - fact_course_daily（集計データシートから日×コース×新規/リピートの件数・金額）
    - fact_cast_monthly（月内集計シートからキャスト×月の集計）
    - dim_cast_seed（スパニストシートからキャストマスタ補完）
    """
    file_path = Path(file_path)
    fname = file_path.name
    m = re.search(r"(\d{4})", fname)
    if not m:
        return {}
    ym = yymm_to_year_month(m.group(1))
    if not ym:
        return {}
    year, month = map(int, ym.split("-"))

    wb = load_workbook(file_path, data_only=True)
    out: dict[str, list] = {
        "fact_daily_sales": [],
        "fact_course_daily": [],
        "fact_cast_monthly": [],
        "dim_cast_seed": [],
    }

    # ---------- 1. 店舗別合計 → fact_daily_sales ----------
    # r3: 三丁目=c1, 上野=c9, 銀座=c16
    # r5: 各店舗ブロックのヘッダ（日付/売上/女子報酬/現金/カード/純利益）
    # r6以降: 日次データ
    if "店舗別合計" in wb.sheetnames:
        ws = wb["店舗別合計"]
        blocks = [(0, 1), (8, 3), (15, 2)]  # (列offset 0-based, store_id) for 三丁目=新宿/上野/銀座
        all_rows = list(ws.iter_rows(min_row=6, max_row=40, values_only=True))
        for off, store_id in blocks:
            for r in all_rows:
                if off >= len(r) or r[off] is None:
                    continue
                day = safe_int(r[off])
                if day is None or day < 1 or day > 31:
                    continue
                try:
                    date_value = pd.Timestamp(year=year, month=month, day=day).date()
                except ValueError:
                    continue
                sales = safe_int(r[off + 1]) if off + 1 < len(r) else None
                reward = safe_int(r[off + 2]) if off + 2 < len(r) else None
                # cash = r[off+3], card = r[off+4]
                profit = safe_int(r[off + 5]) if off + 5 < len(r) else None
                if not sales:
                    continue
                out["fact_daily_sales"].append({
                    "date": date_value,
                    "store_id": store_id,
                    "sales": sales,
                    "case_count": None,
                    "unit_price": None,
                    "gross_profit": profit,
                    "cost": None,
                    "new_count": None,
                    "repeat_count": None,
                    "repeat_rate": None,
                    "source_file": fname,
                })

    # ---------- 2. 集計データ → fact_course_daily ----------
    # r2 ヘッダ：日付 + 「新規60案件数/金額」「新規90」… の対
    if "集計データ" in wb.sheetnames:
        ws = wb["集計データ"]
        rows = list(ws.iter_rows(min_row=2, max_row=400, values_only=True))
        headers = [normalize_text(v) for v in rows[0]]

        # (course_id, new_or_repeat, 案件数ヘッダ, 金額ヘッダ)
        course_map = [
            (1, "new", "新規60案件数", "新規60金額"),
            (2, "new", "新規90案件数", "新規90金額"),
            (3, "new", "新規120案件数", "新規120金額"),
            (4, "new", "新規オイル90案件数", "新規オイル90金額"),
            (5, "new", "新規オイル120案件数", "新規オイル120金額"),
            (1, "repeat", "リピート60案件数", "リピート60金額"),
            (2, "repeat", "リピート90案件数", "リピート90金額"),
            (3, "repeat", "リピート120案件数", "リピート120金額"),
            (4, "repeat", "リピートオイル90案件数", "リピートオイル90金額"),
            (5, "repeat", "リピートオイル120案件", "リピートオイル120金額"),
            (6, "new", "新規クリーム案件数", None),
            (6, "repeat", "リピートクリーム案件数", None),
        ]

        def hdr_idx(name: str) -> Optional[int]:
            try:
                return headers.index(name)
            except ValueError:
                return None

        for r in rows[1:]:
            if not r or r[0] is None:
                continue
            date_value = r[0]
            if hasattr(date_value, "date"):
                date_value = date_value.date()
            elif isinstance(date_value, (int, float)):
                try:
                    date_value = pd.Timestamp(year=year, month=month, day=int(date_value)).date()
                except ValueError:
                    continue
            else:
                continue

            for course_id, nor, case_hdr, amount_hdr in course_map:
                c_idx = hdr_idx(case_hdr) if case_hdr else None
                a_idx = hdr_idx(amount_hdr) if amount_hdr else None
                cnt = safe_int(r[c_idx]) if c_idx is not None and c_idx < len(r) else None
                amt = safe_int(r[a_idx]) if a_idx is not None and a_idx < len(r) else None
                if not cnt and not amt:
                    continue
                # 店舗が日報集計データには無い → store_id=0（全社）を入れて集計用とする
                out["fact_course_daily"].append({
                    "date": date_value,
                    "store_id": 0,
                    "course_id": course_id,
                    "new_or_repeat": nor,
                    "case_count": cnt or 0,
                    "sales": amt or 0,
                    "source_file": fname,
                })

    # ---------- 3. 月内集計 → fact_cast_monthly ----------
    if "月内集計" in wb.sheetnames:
        ws = wb["月内集計"]
        rows = list(ws.iter_rows(min_row=2, max_row=200, values_only=True))
        # r2: スパニスト/案件数/指名/（本指名）/売上/報酬/粗利
        for r in rows[1:]:
            if not r or r[0] is None:
                continue
            cast_name = normalize_cast_name(r[0])
            if not cast_name or cast_name == "スパニスト":
                continue
            case_count = safe_int(r[1]) if len(r) > 1 else 0
            nom = safe_int(r[2]) if len(r) > 2 else 0
            main_nom = safe_int(r[3]) if len(r) > 3 else 0
            sales = safe_int(r[4]) if len(r) > 4 else 0
            reward = safe_int(r[5]) if len(r) > 5 else 0
            gp = safe_int(r[6]) if len(r) > 6 else 0
            if not (sales or case_count):
                continue
            out["fact_cast_monthly"].append({
                "year_month": ym,
                "cast_name": cast_name,
                "sales": sales or 0,
                "reward": reward or 0,
                "gross_profit": gp or 0,
                "case_count": case_count or 0,
                "work_hours": None,
                "store_nomination_rate": (nom / case_count) if case_count else None,
                "store_main_nomination_rate": (main_nom / case_count) if case_count else None,
                "haken_nomination_rate": None,
                "haken_main_nomination_rate": None,
                "survey_score": None,
                "case_count_store": case_count,
                "nomination_store": nom,
                "main_nomination_store": main_nom,
                "sales_store": sales,
                "reward_store": reward,
                "gross_profit_store": gp,
                "total_sales": sales,
                "total_reward": reward,
                "total_gross_profit": gp,
                "total_case_count": case_count,
                "main_nomination_total": main_nom,
                "source_file": fname,
            })

    # ---------- 4. スパニスト → dim_cast_seed ----------
    if "スパニスト" in wb.sheetnames:
        ws = wb["スパニスト"]
        rows = list(ws.iter_rows(min_row=1, max_row=200, values_only=True))
        headers = [normalize_text(v) for v in rows[0]]

        def col(*cands):
            for c in cands:
                if c in headers:
                    return headers.index(c)
            return None

        c_name = col("店舗名", "スパニスト名")
        c_haken = col("派遣名")
        c_priority = col("優先度")
        c_area = col("固定エリア", "エリア")
        c_staff = col("担当スタッフ")
        c_status = col("ステータス", "状態")

        for r in rows[1:]:
            if c_name is None or c_name >= len(r) or not r[c_name]:
                continue
            nm = normalize_cast_name(r[c_name])
            if not nm:
                continue
            prio_raw = normalize_text(r[c_priority]) if c_priority is not None and c_priority < len(r) else None
            from etl.normalize import parse_priority
            out["dim_cast_seed"].append({
                "cast_name": nm,
                "haken_name": normalize_text(r[c_haken]) if c_haken is not None and c_haken < len(r) else None,
                "fixed_area": normalize_text(r[c_area]) if c_area is not None and c_area < len(r) else None,
                "status": normalize_text(r[c_status]) if c_status is not None and c_status < len(r) else None,
                "hire_date": None,
                "priority": parse_priority(r[c_priority]) if c_priority is not None and c_priority < len(r) else None,
                "priority_raw": prio_raw,
            })

    wb.close()

    result: dict[str, pd.DataFrame] = {}
    for k, lst in out.items():
        if lst:
            df = pd.DataFrame(lst)
            if k == "dim_cast_seed":
                df = df.drop_duplicates(subset=["cast_name"], keep="last")
            elif k == "fact_daily_sales":
                df = df.drop_duplicates(subset=["date", "store_id"], keep="last")
            elif k == "fact_course_daily":
                df = df.drop_duplicates(subset=["date", "store_id", "course_id", "new_or_repeat"], keep="last")
            elif k == "fact_cast_monthly":
                df = df.drop_duplicates(subset=["year_month", "cast_name"], keep="last")
            result[k] = df
    return result


WB_DAY_PATTERN = re.compile(r"^(\d{1,2})$")


def load_whiteboard(file_path: Path) -> dict[str, pd.DataFrame]:
    """`ヘッドホワイトボード{YYMM}.xlsx` → fact_attendance, fact_ban_customer。

    勤怠シートは横長ピボット。
    - 行3: 源氏名（2列マージ）
    - 行5〜34: 日次データ。奇数列(C,E,…)=ステータス、偶数列(D,F,…)=稼働時間
    """
    file_path = Path(file_path)
    fname = file_path.name

    # ファイル名から YYMM を抜き出す
    m = re.search(r"(\d{4})", fname)
    if not m:
        return {}
    from etl.normalize import yymm_to_year_month
    ym = yymm_to_year_month(m.group(1))
    if not ym:
        return {}
    year, month = map(int, ym.split("-"))

    wb = load_workbook(file_path, data_only=True, read_only=True)

    # --- 勤怠シート → fact_attendance ---
    attendance_rows: list[dict] = []
    if "勤怠" in wb.sheetnames:
        ws = wb["勤怠"]
        all_rows = list(ws.iter_rows(min_row=1, max_row=40, values_only=True))
        if len(all_rows) >= 5:
            row3 = all_rows[2]  # 源氏名行
            # 奇数列(0,2,4,...の0-indexed)に源氏名が入る
            cast_cols: list[tuple[int, str]] = []
            col_idx = 2  # 0-indexed: C列
            while col_idx < len(row3):
                nm = normalize_cast_name(row3[col_idx]) if row3[col_idx] else None
                if nm:
                    cast_cols.append((col_idx, nm))
                col_idx += 2  # 2列ペア
            # 日次データ（r5〜r34）
            for day_offset, day_row in enumerate(all_rows[4:34]):
                day_num = day_offset + 1
                try:
                    d = pd.Timestamp(year=year, month=month, day=day_num).date()
                except ValueError:
                    continue
                for col_i, cast_name in cast_cols:
                    if col_i >= len(day_row):
                        continue
                    status = normalize_text(day_row[col_i])
                    hours_idx = col_i + 1
                    hours = safe_float(day_row[hours_idx]) if hours_idx < len(day_row) else None
                    if not status and not hours:
                        continue
                    attendance_rows.append({
                        "date": d,
                        "cast_name": cast_name,
                        "status": status or "未記入",
                        "work_hours": hours,
                        "source_file": fname,
                    })

    # ※ 出禁リストはダッシュボード対象外（ユーザー指示で削除済み）
    wb.close()
    out: dict[str, pd.DataFrame] = {}
    if attendance_rows:
        out["fact_attendance"] = pd.DataFrame(attendance_rows)
    return out


def load_training(file_path: Path) -> dict[str, pd.DataFrame]:
    """`ヘッド研修日程表.xlsx` → fact_training（研修状況共有シートから）。"""
    file_path = Path(file_path)
    fname = file_path.name
    wb = load_workbook(file_path, data_only=True, read_only=True)
    if "研修状況共有" not in wb.sheetnames:
        wb.close()
        return {}

    ws = wb["研修状況共有"]
    rows = list(ws.iter_rows(min_row=1, max_row=200, values_only=True))
    if len(rows) < 2:
        wb.close()
        return {}

    headers = [normalize_text(v) if v else "" for v in rows[0]]
    # 想定列: 名前/源氏名 + 研修種別1..N
    # 「店舗名」列名 = 源氏名（既存知見）
    name_col = None
    for cand in ("店舗名", "スパニスト", "名前", "源氏名"):
        if cand in headers:
            name_col = headers.index(cand)
            break
    if name_col is None:
        wb.close()
        return {}

    # 研修種別列：名前列より右の非空ヘッダ
    training_cols = [
        (i, h) for i, h in enumerate(headers)
        if i > name_col and h and h not in ("休業中",)
    ]

    fact_rows = []
    for r in rows[1:]:
        if name_col >= len(r) or not r[name_col]:
            continue
        cast_name = normalize_cast_name(r[name_col])
        if not cast_name:
            continue
        for i, tname in training_cols:
            if i >= len(r):
                continue
            val = r[i]
            status = "未着手"
            completed_date = None
            if val is None or (isinstance(val, str) and val.strip() == ""):
                continue
            if hasattr(val, "date"):
                status = "済"
                completed_date = val.date() if hasattr(val, "date") else val
            else:
                v = normalize_text(val)
                if v == "済":
                    status = "済"
                elif v in ("研修中", "1回目", "日程調整中", "休業中"):
                    status = v
                else:
                    status = v[:20]
            fact_rows.append({
                "cast_name": cast_name,
                "training_type": tname,
                "status": status,
                "completed_date": completed_date,
                "source_file": fname,
            })

    wb.close()
    if fact_rows:
        return {"fact_training": pd.DataFrame(fact_rows)}
    return {}


# 自動判定ディスパッチ
LOADER_BY_FILENAME_PATTERN: list[tuple[re.Pattern, callable]] = [
    (re.compile(r"ヘッド店舗KPI"), load_store_kpi),
    (re.compile(r"ヘッドスパニスト一覧管理表"), load_cast_master),
    (re.compile(r"aoスパニスト評価"), load_cast_evaluation),
    (re.compile(r"ao合計日報"), load_daily_report),
    (re.compile(r"ヘッドホワイトボード"), load_whiteboard),
    (re.compile(r"ヘッド研修日程表"), load_training),
]


def dispatch(file_path: Path) -> Optional[callable]:
    name = Path(file_path).name
    for pattern, loader in LOADER_BY_FILENAME_PATTERN:
        if pattern.search(name):
            return loader
    return None
